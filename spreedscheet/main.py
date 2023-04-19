import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
product_per_supplier = {}
total_inv_value_per_supplier = {}
products_under_10_inv = {}


for row in range(2, product_list.max_row+1):
    supplier = product_list.cell(row,4).value
    inventory = product_list.cell(row,2).value
    price = product_list.cell(row,3).value
    product_num = product_list.cell(row,1).value
    inventory_price = product_list.cell(row,5)

    # calculation number of product per supplier
    if supplier in product_per_supplier:
        product_count = product_per_supplier[supplier]
        product_per_supplier[supplier] = product_count + 1
    else:
        product_per_supplier[supplier] = 1

    #  calculation total inventory value per supplier  
    if supplier in total_inv_value_per_supplier:
        inv_value = total_inv_value_per_supplier.get(supplier) 
        total_inv_value_per_supplier[supplier] =  inv_value + inventory*price 
    else:
        total_inv_value_per_supplier[supplier] = inventory*price   

    #get products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[product_num]=inventory

    #add column for inventory price 
    inventory_price.value = inventory*price
product_list.cell(1,5).value = "INV*P"




print(f'The number of product per supplier:-\n{product_per_supplier}')
print(f'The total inventory value per supplier:-\n{total_inv_value_per_supplier}')
print(f'products with inventory less than 10:-\n{ products_under_10_inv}')
inv_file.save("modified_inventory.xlsx")








